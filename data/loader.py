"""Data loading and preprocessing functions."""

from pathlib import Path
import re
from typing import Dict, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from app_constants import ColumnNames, StockColumnNames, StockSheetNames


# Directory configuration
PROJECT_ROOT = Path(__file__).parent.parent
DATA_DIR = PROJECT_ROOT / 'data'
RAW_DATA_DIR = DATA_DIR / 'raw'
DEFAULT_BUDGET_FILENAME = "budgets.csv"
EMPTY_DF = pd.DataFrame()
PAYOUT_WORKBOOK_TAXABLE_ACCOUNT_ALIAS = {
    "IKBR": "Interactive Brokers",
    "Robinhood": "Robinhood",
    "Zerodha": "Zerodha",
    "Fidelity-Ind": "Fidelity Brokerage",
    "Fidelity-Roth": "Fidelity Roth",
}


def _resolve_raw_path(filename: str) -> Path:
    """Return a path inside the raw data directory."""
    return RAW_DATA_DIR / filename


def _resolve_data_path(filename: str) -> Path:
    """Return a path inside the data directory."""
    return DATA_DIR / filename


def _show_missing_file_error(filepath: Path, context: str) -> None:
    """Display a consistent missing-file message with recovery guidance."""
    st.error(f"{context} file not found.")
    st.caption(f"Expected path: `{filepath}`")
    st.info("Add the file in the expected location or update the loader configuration before retrying.")


def _show_load_error(filepath: Path, context: str, error: Exception) -> None:
    """Display a consistent data-load failure message."""
    st.error(f"Unable to load {context.lower()}.")
    st.caption(f"Source: `{filepath}`")
    st.info(f"Check the file format, required columns, and sheet names. Details: {error}")


def _coerce_float(value: object, fallback: float = 0.0) -> float:
    """Safely coerce workbook values to float."""
    try:
        if value is None or pd.isna(value):
            return fallback
        return float(value)
    except (TypeError, ValueError):
        return fallback


def _extract_formula_numbers(formula: object) -> list[float]:
    """Extract numeric literals from an Excel formula string."""
    return [float(match) for match in re.findall(r"(?<![A-Za-z])\d+(?:\.\d+)?", str(formula or ""))]


def _mask_account_identifier(value: object) -> str:
    """Return a short, user-friendly masked identifier."""
    identifier = str(value).strip()
    if not identifier or identifier.lower() == "nan":
        return ""

    suffix = identifier[-4:] if len(identifier) >= 4 else identifier
    return f"{identifier}"


def _join_account_parts(parts: list[str]) -> str:
    """Join non-empty account label parts with a consistent separator."""
    clean_parts = [part for part in parts if part]
    return " | ".join(clean_parts)


def _build_networth_account_identity(data: pd.DataFrame) -> pd.DataFrame:
    """Create stable account key/display columns for net worth data."""
    normalized = data.copy()

    alias_map = {
        "account type": ColumnNames.ACCOUNT_TYPE,
        "account_type": ColumnNames.ACCOUNT_TYPE,
        "account subtype": ColumnNames.CATEGORY,
        "account_subtype": ColumnNames.CATEGORY,
        "financial institution": ColumnNames.INSTITUTION,
        "financial_institution": ColumnNames.INSTITUTION,
        "institution": ColumnNames.INSTITUTION,
        "brokerage": ColumnNames.INSTITUTION,
        "as of date": ColumnNames.MONTH,
        "as_of_date": ColumnNames.MONTH,
        "balance": ColumnNames.AMOUNT,
        "account id": ColumnNames.ACCOUNT_ID,
        "account_id": ColumnNames.ACCOUNT_ID,
        "account number": ColumnNames.ACCOUNT_ID,
        "account_number": ColumnNames.ACCOUNT_ID,
        "acct number": ColumnNames.ACCOUNT_ID,
        "acct_number": ColumnNames.ACCOUNT_ID,
        "acct id": ColumnNames.ACCOUNT_ID,
        "acct_id": ColumnNames.ACCOUNT_ID,
    }

    rename_map = {}
    for column in normalized.columns:
        normalized_name = str(column).strip().lower()
        if normalized_name in alias_map and alias_map[normalized_name] not in normalized.columns:
            rename_map[column] = alias_map[normalized_name]

    if rename_map:
        normalized = normalized.rename(columns=rename_map)

    if ColumnNames.ACCOUNT not in normalized.columns:
        if ColumnNames.CATEGORY in normalized.columns:
            normalized[ColumnNames.ACCOUNT] = normalized[ColumnNames.CATEGORY]
        elif ColumnNames.INSTITUTION in normalized.columns:
            normalized[ColumnNames.ACCOUNT] = normalized[ColumnNames.INSTITUTION]
        else:
            normalized[ColumnNames.ACCOUNT] = "Unknown Account"

    account_series = normalized[ColumnNames.ACCOUNT].fillna("Unknown Account").astype(str).str.strip()
    normalized[ColumnNames.ACCOUNT] = account_series

    if ColumnNames.ACCOUNT_ID in normalized.columns:
        account_id_series = normalized[ColumnNames.ACCOUNT_ID].fillna("").astype(str).str.strip()
    else:
        account_id_series = pd.Series("", index=normalized.index, dtype="object")

    institution_series = (
        normalized[ColumnNames.INSTITUTION].fillna("").astype(str).str.strip()
        if ColumnNames.INSTITUTION in normalized.columns
        else pd.Series("", index=normalized.index, dtype="object")
    )
    masked_ids = account_id_series.apply(_mask_account_identifier)

    if (account_id_series != "").any():
        key_suffix = institution_series.where(institution_series != "", masked_ids)
        normalized[ColumnNames.ACCOUNT_KEY] = account_series.where(
            account_id_series == "",
            account_series + "::" + key_suffix + "::" + account_id_series,
        )
        normalized[ColumnNames.ACCOUNT_DISPLAY] = [
            _join_account_parts([account_name, institution, masked_id])
            for account_name, institution, masked_id in zip(
                account_series,
                institution_series,
                masked_ids,
            )
        ]
    elif (institution_series != "").any():
        normalized[ColumnNames.ACCOUNT_KEY] = [
            _join_account_parts([account_name, institution])
            for account_name, institution in zip(account_series, institution_series)
        ]
        normalized[ColumnNames.ACCOUNT_DISPLAY] = normalized[ColumnNames.ACCOUNT_KEY]
    else:
        normalized[ColumnNames.ACCOUNT_KEY] = account_series
        normalized[ColumnNames.ACCOUNT_DISPLAY] = account_series

    if ColumnNames.CATEGORY in normalized.columns:
        normalized[ColumnNames.CATEGORY] = (
            normalized[ColumnNames.CATEGORY].fillna("Unknown").astype(str).str.strip()
        )
    if ColumnNames.ACCOUNT_TYPE in normalized.columns:
        normalized[ColumnNames.ACCOUNT_TYPE] = (
            normalized[ColumnNames.ACCOUNT_TYPE].fillna("Unknown").astype(str).str.strip()
        )
    if ColumnNames.INSTITUTION in normalized.columns:
        normalized[ColumnNames.INSTITUTION] = institution_series
    if ColumnNames.ACCOUNT_ID in normalized.columns:
        normalized[ColumnNames.ACCOUNT_ID] = account_id_series

    return normalized


def _load_excel_sheet(filepath: Path, sheet_name: str, context: str) -> pd.DataFrame:
    """Load a single Excel sheet with consistent error handling."""
    try:
        return pd.read_excel(filepath, sheet_name=sheet_name)
    except FileNotFoundError:
        _show_missing_file_error(filepath, context)
    except ValueError as error:
        st.error(f"Required sheet `{sheet_name}` not found for {context.lower()}.")
        st.caption(f"Source: `{filepath}`")
        st.info(f"Add the missing sheet and retry. Details: {error}")
    except Exception as error:
        _show_load_error(filepath, context, error)

    return EMPTY_DF.copy()


@st.cache_data
def load_payout_sheet_defaults(filename: str = "Investment.xlsx") -> dict[str, object]:
    """Load workbook-backed payout assumptions from the payout sheet."""
    workbook_defaults: dict[str, object] = {
        "available": False,
        "taxable_profit_assumptions": {},
        "annual_salary": None,
        "vacation_hours": None,
        "vacation_after_tax_factor": 0.73,
        "usd_to_inr_rate": 93.95,
        "capital_gains_tax_rate": 0.15,
        "vacation_payout": None,
        "missing_fields": [],
        "source": str(_resolve_raw_path(filename)),
    }

    workbook_path = _resolve_raw_path(filename)
    try:
        values_wb = load_workbook(workbook_path, data_only=True, read_only=True)
        formulas_wb = load_workbook(workbook_path, data_only=False, read_only=True)
        values_sheet = values_wb["payout"]
        formulas_sheet = formulas_wb["payout"]

        workbook_profit_assumptions: dict[str, float] = {}
        for row_idx in range(3, 8):
            workbook_label = str(values_sheet[f"B{row_idx}"].value or "").strip()
            account_label = PAYOUT_WORKBOOK_TAXABLE_ACCOUNT_ALIAS.get(workbook_label)
            if not account_label:
                continue
            workbook_profit_assumptions[account_label] = _coerce_float(values_sheet[f"D{row_idx}"].value)

        if workbook_profit_assumptions:
            workbook_defaults["taxable_profit_assumptions"] = workbook_profit_assumptions

        salary_formula_numbers = _extract_formula_numbers(formulas_sheet["G3"].value)
        if salary_formula_numbers:
            workbook_defaults["annual_salary"] = salary_formula_numbers[0]

        vacation_formula_numbers = _extract_formula_numbers(formulas_sheet["G4"].value)
        factor_candidates = [number for number in vacation_formula_numbers if 0 < number < 1]
        hour_candidates = [number for number in vacation_formula_numbers if number >= 10]
        if factor_candidates:
            workbook_defaults["vacation_after_tax_factor"] = factor_candidates[-1]
        if hour_candidates:
            workbook_defaults["vacation_hours"] = hour_candidates[0]

        capital_gains_formula_numbers = _extract_formula_numbers(formulas_sheet["D8"].value)
        if capital_gains_formula_numbers:
            workbook_defaults["capital_gains_tax_rate"] = max(0.0, 1 - capital_gains_formula_numbers[-1])

        workbook_fx_rate = _coerce_float(values_sheet["E12"].value, fallback=93.95)
        if workbook_fx_rate > 0:
            workbook_defaults["usd_to_inr_rate"] = workbook_fx_rate

        workbook_vacation_payout = _coerce_float(values_sheet["G4"].value, fallback=0.0)
        if workbook_vacation_payout > 0:
            workbook_defaults["vacation_payout"] = workbook_vacation_payout

        values_wb.close()
        formulas_wb.close()
    except Exception as error:
        workbook_defaults["error"] = str(error)
        return workbook_defaults

    missing_fields: list[str] = []
    if not workbook_defaults["taxable_profit_assumptions"]:
        missing_fields.append("taxable profit assumptions")
    if workbook_defaults["annual_salary"] is None:
        missing_fields.append("annual salary seed")
    if workbook_defaults["vacation_hours"] is None:
        missing_fields.append("vacation hours")

    workbook_defaults["missing_fields"] = missing_fields
    workbook_defaults["available"] = not missing_fields
    return workbook_defaults


def _normalize_stock_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize stock sheet column names to a predictable canonical shape."""
    if df.empty:
        return df.copy()

    normalized_df = df.copy()
    normalized_df.columns = [str(col).strip() for col in normalized_df.columns]

    normalized_map = {
        "date": StockColumnNames.DATE,
        "ticker": StockColumnNames.TICKER,
        "symbol": StockColumnNames.TICKER,
        "quantity": StockColumnNames.QUANTITY,
        "brokerage": StockColumnNames.BROKERAGE,
        "account name": StockColumnNames.ACCOUNT_NAME,
        "investment type": StockColumnNames.INVESTMENT_TYPE,
    }

    rename_map = {}
    for col in normalized_df.columns:
        normalized_key = col.lower().strip()
        if normalized_key in normalized_map:
            rename_map[col] = normalized_map[normalized_key]

    return normalized_df.rename(columns=rename_map)


@st.cache_data
def load_networth_data(filename: str = "Networth.csv") -> pd.DataFrame:
    """Load and preprocess net worth data from CSV.
    
    Args:
        filename: Name of CSV file to load from raw data directory
        
    Returns:
        Preprocessed DataFrame with datetime month column and formatted strings
    """
    filepath = _resolve_raw_path(filename)
    fallback_filepath = _resolve_raw_path("Investment.xlsx")
    fallback_sheet = "long_data"
    source_filepath = filepath
    
    try:
        data = pd.read_csv(filepath)
    except FileNotFoundError:
        source_filepath = fallback_filepath
        data = _load_excel_sheet(fallback_filepath, fallback_sheet, "Net worth data")
        if data.empty:
            _show_missing_file_error(filepath, "Net worth data")
            return EMPTY_DF.copy()
    except Exception as e:
        _show_load_error(filepath, "Net worth data", e)
        return EMPTY_DF.copy()

    try:
        data = _build_networth_account_identity(data)

        # Process date and amount columns
        data[ColumnNames.MONTH] = pd.to_datetime(data[ColumnNames.MONTH])
        data[ColumnNames.AMOUNT] = data[ColumnNames.AMOUNT].round().astype(int)
        data[ColumnNames.MONTH_STR] = data[ColumnNames.MONTH].dt.strftime('%b-%Y')
        data = data.sort_values(ColumnNames.MONTH)

        return data
    except Exception as e:
        _show_load_error(source_filepath, "Net worth data", e)
        return EMPTY_DF.copy()


@st.cache_data
def load_expense_transactions(filename: str = 'transactions.xlsx') -> pd.DataFrame:
    """Load transaction data from CSV file.
    
    Args:
        filename: Name of CSV file to load from raw data directory
        
    Returns:
        DataFrame with transaction data including date and type columns
    """
    filepath = _resolve_raw_path(filename)
    
    try:
        df = pd.read_excel(filepath)
        df[ColumnNames.DATE] = pd.to_datetime(df[ColumnNames.DATE])
        
        # Add 'type' column if it doesn't exist
        if 'type' not in df.columns:
            df['type'] = 'expense'
        
        return df
        
    except FileNotFoundError:
        _show_missing_file_error(filepath, "Expense transactions")
        return EMPTY_DF.copy()
    except Exception as e:
        _show_load_error(filepath, "Expense transactions", e)
        return EMPTY_DF.copy()


@st.cache_data
def load_budgets(filename: str = DEFAULT_BUDGET_FILENAME) -> Dict[str, float]:
    """Load budget data from CSV or Excel file.
    
    Args:
        filename: Name of budget file (supports .csv or .xlsx)
        
    Returns:
        Dictionary mapping category names to budget amounts
    """
    csv_path = _resolve_data_path(filename)
    xlsx_name = Path(filename).with_suffix('.xlsx').name
    xlsx_path = _resolve_data_path(xlsx_name)
    
    # Try CSV first
    if csv_path.exists():
        try:
            budget_df = pd.read_csv(csv_path)
            return dict(zip(budget_df[ColumnNames.DATE], budget_df['budget']))
        except Exception as e:
            st.warning("Budget file could not be loaded from CSV. Using default budget values.")
            st.caption(f"Source: `{csv_path}`")
            st.info(f"Check for `date` and `budget` columns. Details: {e}")
    
    # Try Excel if CSV not found or failed
    elif xlsx_path.exists():
        try:
            budget_df = pd.read_excel(xlsx_path)
            return dict(zip(budget_df[ColumnNames.DATE], budget_df['budget']))
        except Exception as e:
            st.warning("Budget file could not be loaded from Excel. Using default budget values.")
            st.caption(f"Source: `{xlsx_path}`")
            st.info(f"Check for `date` and `budget` columns. Details: {e}")
    
    # Return defaults if no file found
    return _get_default_budgets()


@st.cache_data
def load_stock_data(filename: str = 'stock_positions.xlsx') -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Load stock trading and historical sheets from the Excel file."""
    file_path = _resolve_raw_path(filename)

    trading_log = _normalize_stock_columns(
        _load_excel_sheet(file_path, StockSheetNames.TRADING_LOG, "Stock tracker data")
    )
    historical = _normalize_stock_columns(
        _load_excel_sheet(file_path, StockSheetNames.HISTORICAL_TRACKING, "Stock tracker data")
    )

    if not trading_log.empty and StockColumnNames.DATE in trading_log.columns:
        trading_log[StockColumnNames.DATE] = pd.to_datetime(trading_log[StockColumnNames.DATE])

    if not historical.empty and StockColumnNames.DATE in historical.columns:
        historical[StockColumnNames.DATE] = pd.to_datetime(historical[StockColumnNames.DATE])

    return trading_log, historical


# if you don't want to provide a file for budget set it here with whatever categories you have
def _get_default_budgets() -> Dict[str, float]:
    """Return default budget values.
    
    Returns:
        Dictionary of default category budgets
    """
    return {
        'Housing': 1200.00,
        'Utilities': 150.00,
        'Food & Dining': 350.00,
        'Entertainment': 100.00,
        'Transportation': 200.00,
        'Miscellaneous': 200.00,
    }
